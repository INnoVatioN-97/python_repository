from openpyxl import Workbook

sheet1 = '오우씥1'
sheet2 = '오우씥2'

wb = Workbook()
ws = wb.create_sheet()
ws.title = '오우씥ㅌㅌㅌㅌㅌ'
ws.sheet_properties.tabColor = 'ff00ff'

ws1 = wb.create_sheet(sheet1)
ws2 = wb.create_sheet(sheet2)

print(wb[sheet1].title)
new_ws = wb[sheet1]
print(new_ws.title)

print(wb.sheetnames)

for i in range(1, 21):
    for j in range(1, 41):
        # print(i)
        ws1.cell(row=i+1, column=j+1).value = i*j
        # ws1.git


wb.save('fucking_excel.xlsx')
